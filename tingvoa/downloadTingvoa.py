#!/usr/bin/env python
#! encoding=utf-8

# Author        : kesalin@gmail.com
# Blog          : http://luozhaohui.github.io
# Date          : 2017/10/20
# Description   : download resources from tingvoa.com 
# Version       : 1.0.0.0
# Python Version: Python 2.7.3
#
# store to excel nees to install openpyxl: pip install openpyxl

import os
import threading
from multiprocessing.dummy import Pool as ThreadPool
import time
import datetime
import re
import string
import urllib2
import timeit
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

#=============================================================================
# request & response
#=============================================================================

# 获取 url 内容
gUseCookie = False
gHeaders = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36',
    'Cookie': 'Put your cookie here'
}

def getHtml(url):
    try :
        if gUseCookie:
            opener = urllib2.build_opener()
            for k, v in gHeaders.items():
                opener.addheaders.append((k, v))
            response = opener.open(url)
            data = response.read().decode('utf-8')
        else:
            request = urllib2.Request(url, None, gHeaders['User-Agent'])
            response = urllib2.urlopen(request)
            data = response.read().decode('utf-8')
    except urllib2.URLError, e :
        if hasattr(e, "code"):
            print "The server couldn't fulfill the request: " + url
            print "Error code: %s" % e.code
        elif hasattr(e, "reason"):
            print "We failed to reach a server. Please check your url: " + url + ", and read the Reason."
            print "Reason: %s" % e.reason
    return data


#=============================================================================
# classes
#=============================================================================
# 
class LevelInfo:
    def __init__(self, name, url):
        self.name = name
        self.url = url
        self.bookInfos = []

    def __str__( self ):
        return "Level:{0}, url:{1}, books:{2}\n".format(self.name, self.url, len(self.bookInfos))
# 
class BookInfo:
    def __init__(self, name, url):
        self.name = name
        self.url = url
        self.chapterInfos = []

    def __str__( self ):
        return "Book:{0}, url:{1}, chapters:[2]".format(self.name, self.url, len(self.chapterInfos))
# 
class ChapterInfo:
    def __init__(self, name, url, mp3):
        self.name = name
        self.url = url
        self.mp3Url = mp3

    def __str__( self ):
        return "Chapter:{0}, url:{1}, mp3:{2}".format(self.name, self.url, self.mp3Url)

# 
class DownloadInfo:
    def __init__(self, name, path, mp3):
        self.name = name
        self.path = path
        self.mp3Url = mp3

    def __str__( self ):
        return "Download:path:{0}, mp3:{1}".format(self.path, self.mp3Url)

#=============================================================================
# utilities
#=============================================================================

def get_root_url(url):
    rootUrl = url
    start = url.find(".")
    if start > 0:
        index = url.find("/", start);
        if index > 0:
            rootUrl = url[0:index]
    return rootUrl


def slow_down():
    time.sleep(0.5)         # slow down a little


def mkdir(path):
    if False == os.path.exists(path):
        os.mkdir(path)


def get_postfix(resUrl):
    index = resUrl.rfind('.')
    if index > 0:
        return resUrl[index:]
    return ""


def get_cpu_count():
    """Return the number of logical CPUs in the system."""
    try:
        return os.sysconf("SC_NPROCESSORS_ONLN")
    except ValueError:
        # mimic os.cpu_count() behavior
        return None


#=============================================================================
# parse
#=============================================================================
def get_book_infos(levelInfo):
    name = levelInfo.name
    url = levelInfo.url
    rootUrl = get_root_url(url)
    #print levelInfo

    slow_down()
    page = getHtml(url)
    soup = BeautifulSoup(page, 'html.parser')

    bookInfos = []
    # get book infos
    mainleftlist = soup.find(id="mainleftlist");
    if (mainleftlist != None):
        #print mainleftlist.prettify()
        leftTitles = mainleftlist.find_all("div", "leftTitle")
        #print "find %d books" % len(leftTitles)
        for leftTitle in leftTitles:
            #print leftTitle.prettify()
            link = leftTitle.find("a");
            if link != None:
                bookName = ""
                if link.string != None:
                    bookName = link.string.strip().encode('utf-8')
                href = link['href'].encode('utf-8')
                bookUrl = rootUrl + href;

                bookInfo = BookInfo(bookName, bookUrl)
                bookInfos.append(bookInfo)

                # get chapter infos
                chapterInfos = []
                subleftList = leftTitle.findNext("div")
                if subleftList != None:
                    #print subleftList.prettify()
                    allLinks = subleftList.findAll("a")
                    for link in allLinks[1:]:
                        chapterName = ""
                        if link.string != None:
                            chapterName = link.string.strip().encode('utf-8')
                        href = link['href'].encode('utf-8')
                        chapterUrl = rootUrl + href;
                        mp3Url = get_mp3_url(chapterUrl)
                        chapterInfo = ChapterInfo(chapterName, chapterUrl, mp3Url)
                        chapterInfos.append(chapterInfo)
                        #print chapterInfo

                bookInfo.chapterInfos = chapterInfos
                #print bookInfo

    return bookInfos


def get_mp3_url(chapterUrl):
    rootUrl = get_root_url(chapterUrl)

    slow_down()
    page = getHtml(chapterUrl)
    soup = BeautifulSoup(page, 'html.parser')
    content = soup.prettify().encode('utf-8')

    mp3Url = ""
    #var mp3 ="Sound/shuchong/aqyjq/tingvoa.com_1.mp3";
    pattern = re.compile(r'(var mp3)(\s*)(=)(\s*)(")(.*)(.mp3)(")')
    match = pattern.search(content)
    if match:
        mp3 = "{0}{1}".format(match.group(6), match.group(7))
        mp3Url = "http://x8.tingvoa.com/{1}".format(rootUrl, mp3)
        #print mp3Url
    return mp3Url


def store_resource(title, levelInfos):
    mkdir(title)

    # store as markdown
    fileName = "{0}.md".format(title)
    path = os.path.join(title, fileName)
    if os.path.isfile(path):
        os.remove(path)

    today = datetime.datetime.now()
    todayStr = today.strftime('%Y-%m-%d %H:%M:%S %z')
    file = open(path, 'a')
    file.write('## {0}\n'.format(title))
    file.write('### 总计 {0} levels\n'.format(len(levelInfos)))
    file.write('### 更新时间: {0}\n'.format(todayStr))

    for i, level in enumerate(levelInfos):
        file.write('\n### Level.{0:d} {1}\n'.format(i + 1, level.name))

        for j, book in enumerate(level.bookInfos):
            file.write('\n#### Book.{0:d} {1}\n'.format(j + 1, book.name))

            for k, chapter in enumerate(book.chapterInfos):
                file.write('##### Chapter.{0:d} {1}, {2}\n'.format(k + 1, chapter.name, chapter.mp3Url))

    file.close()
    return path


def store_to_excel(resPath):
    workbook = Workbook()
    resBasename = os.path.basename(resPath);
    sheetName = os.path.splitext(resBasename)[0]
    workSheet = workbook.create_sheet(0, sheetName.decode('utf-8'))

    # for i, sheetName in enumerate(workbook.get_sheet_names()):
    #     print "sheet {0}: {1}".format(i, sheetName)

    rootDir = os.path.dirname(resPath)
    currentLevelDir = ""
    currentBookDir = ""
    with open(resPath, "r") as file:
        for line in file:
            ##### Chapter.1 爱情与金钱：1 Chapter, http://x8.tingvoa.com/Sound/shuchong/aqyjq/tingvoa.com_1.mp3
            pattern = re.compile(r'(#*)(\s*Chapter\.[0-9]*)(.*)')
            match = pattern.search(line)
            if match:
                info = match.group(3).strip()
                items = info.split(',')
                if len(items) >= 2:
                    name = items[0]
                    mp3Url = items[1]
                    path = os.path.join(currentBookDir, "{0}{1}".format(name, get_postfix(mp3Url)))
                    workSheet.append([name, path, mp3Url]);
                continue

            #### Book.1 爱情与金钱
            pattern = re.compile(r'(#*)(\s*Book\.[0-9]*)(.*)')
            match = pattern.search(line)
            if match:
                name = match.group(3).strip()
                currentBookDir = os.path.join(currentLevelDir, name)

                info = "{0}{1}".format(match.group(2), match.group(3))
                workSheet.append([""])
                workSheet.append([info])

                print ">> current book: {0}".format(currentBookDir)
                continue

            ### Level.1 书虫第一级
            pattern = re.compile(r'(#*)(\s*Level\.[0-9]*)(.*)')
            match = pattern.search(line)
            if match:
                name = match.group(3).strip()
                currentLevelDir = os.path.join(rootDir, name)

                info = "{0}{1}".format(match.group(2), match.group(3))
                workSheet.append([""])
                workSheet.append([info.strip()])

                print ">> current level: {0}".format(currentLevelDir)
                continue

    # Save the file
    excelFile = "{0}.xlsx".format(sheetName)
    excelPath = os.path.join(os.path.dirname(resPath), excelFile)
    if os.path.isfile(excelPath):
        os.remove(excelPath)
    workbook.save(excelPath)


def print_level_infos(levelInfos):
    for i, level in enumerate(levelInfos):
        print '\n### Level.{0:d} {1}\n'.format(i + 1, level.name)

        for j, book in enumerate(level.bookInfos):
            print '\n#### Book.{0:d} {1}\n'.format(j + 1, book.name)

            for k, chapter in enumerate(book.chapterInfos):
                print '##### Chapter.{0:d} {1}, {2}\n'.format(k + 1, chapter.name, chapter.mp3Url)


# parse resource url
def parse(url):
    start = timeit.default_timer()

    rootUrl = get_root_url(url)
    print "rootUrl: " + rootUrl

    levelInfos = []

    page = getHtml(url)
    soup = BeautifulSoup(page, 'html.parser')

    # get page title
    #pageTitle = soup.html.head.title.string.encode('utf-8')
    #print " > page title：" + pageTitle

    # get resource title
    resourceTitle = ""
    content = soup.find(id="containertow")
    if content != None:
        titleContent = content.find("div", "catmenutitle");
        if titleContent != None and titleContent.string != None:
            resourceTitle = titleContent.string.strip().encode('utf-8')
            #print " > resource title：" + resourceTitle

    # get level infos
    mainleftlist = soup.find(id="mainleftlist");
    if (mainleftlist != None):
        #print mainleftlist.prettify()
        leftTitles = mainleftlist.find_all("div", "leftTitle")
        print "find %d levels" % len(leftTitles)
        for leftTitle in leftTitles:
            #print leftTitle.prettify()
            link = leftTitle.find("a");
            if link != None:
                levelName = ""
                if link.string != None:
                    levelName = link.string.strip().encode('utf-8')
                href = link['href'].encode('utf-8')
                levelUrl = rootUrl + href;

                levelInfo = LevelInfo(levelName, levelUrl)
                levelInfos.append(levelInfo)
                print levelInfo

    for levelInfo in levelInfos:
        levelInfo.bookInfos = get_book_infos(levelInfo)

    print_level_infos(levelInfos)

    path = store_resource(resourceTitle, levelInfos)

    download_resource(path)

    store_to_excel(path)

    elapsed = timeit.default_timer() - start
    print " > 下载完成，耗时 {0} 秒".format(elapsed)

#=============================================================================
# download
#=============================================================================
def download_resource(resPath):
    downloadInfos = []

    rootDir = os.path.dirname(resPath)
    currentLevelDir = ""
    currentBookDir = ""
    with open(resPath, "r") as file:
        for line in file:
            ##### Chapter.1 爱情与金钱：1 Chapter, http://x8.tingvoa.com/Sound/shuchong/aqyjq/tingvoa.com_1.mp3
            pattern = re.compile(r'(#*)(\s*Chapter\.[0-9]*)(.*)')
            match = pattern.search(line)
            if match:
                info = match.group(3).strip()
                items = info.split(',')
                if len(items) >= 2:
                    name = items[0]
                    mp3Url = items[1]
                    path = os.path.join(currentBookDir, "{0}{1}".format(name, get_postfix(mp3Url)))

                    downloadInfo = DownloadInfo(name, path, mp3Url)
                    downloadInfos.append(downloadInfo)
                    #print ">> current chapter: {0}".format(downloadInfo)
                continue

            #### Book.1 爱情与金钱
            pattern = re.compile(r'(#*)(\s*Book\.[0-9]*)(.*)')
            match = pattern.search(line)
            if match:
                name = match.group(3).strip()
                currentBookDir = os.path.join(currentLevelDir, name)
                mkdir(currentBookDir)
                print ">> current book: {0}".format(currentBookDir)
                continue

            ### Level.1 书虫第一级
            pattern = re.compile(r'(#*)(\s*Level\.[0-9]*)(.*)')
            match = pattern.search(line)
            if match:
                name = match.group(3).strip()
                currentLevelDir = os.path.join(rootDir, name)
                mkdir(currentLevelDir)
                print ">> current level: {0}".format(currentLevelDir)
                continue

    download(downloadInfos, get_cpu_count())

def download(infos, threads=2):
    pool = ThreadPool(threads)
    pool.map(download_info, infos)
    pool.close()
    pool.join()
    print ">> download parllel done!"

def download_info(info):
    print "> download {0} to {1}".format(info.mp3Url, info.path)
    remote = urllib2.urlopen(info.mp3Url) 
    with open(info.path, "wb") as local:
       local.write(remote.read()) 

#=============================================================================
# main
#=============================================================================
gResourceUrl = "http://www.tingvoa.com/bookworm/"

if __name__ == '__main__': 
    parse(gResourceUrl)

    #download_resource("书虫牛津英语读物/书虫牛津英语读物.md")
    #store_to_excel("书虫牛津英语读物/书虫牛津英语读物.md")